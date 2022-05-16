import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties

# 获取当前文件的上级路径
this_path = Path(__file__).resolve().parent

# 读取该路径下的所有Excel文件
df_list = []
for file in (this_path / 'data/sales_data').rglob('*.xls*'):
    df = pd.read_excel(file)
    # 将df添加到列表中
    df_list.append(df)

# 生成DataFrame
df = pd.concat(df_list)

# 透视汇总数据(使用日期汇总每个门店销售总金额)
pivot = pd.pivot_table(data=df,
                       values='amount',
                       index='transaction_date',
                       columns='store',
                       aggfunc='sum')

# 采样得到月级每个门店销售总金额
summary = pivot.resample('M').sum()
summary.index.name = 'Month'

# 为数据添加行汇总
summary = summary.append(summary.sum(axis=0).rename('Total'))
# 为数据添加列汇总
summary['Total'] = summary.sum(axis=1)

# 按商户销售量排名
summary = summary.loc[:, summary.sum().sort_values().index]

# 设置写入起始位置
start_row, start_col = 3, 2
# 写入数据的维度
nrows, ncols = summary.shape

with pd.ExcelWriter(this_path / 'data/output/sales_report_1.xlsx', engine='openpyxl') as writer:
    summary.to_excel(writer,
                     sheet_name='sales_report',
                     startrow=start_row - 1,  # 写出到目标Excel文件(ExcelWriter写入索引从0开始)
                     startcol=start_col - 1)
    # 获取工作簿对象
    book = writer.book
    # 获取sheet对象
    sheet = writer.sheets['sales_report']
    # 设置表头（cell索引是从1开始的）
    sheet.cell(row=1, column=start_col, value='Sales Report')
    sheet.cell(row=1, column=start_col).font = Font(color='0000FF', size=14, bold=True)
    # 设置数据域格式
    for i in range(start_row + 1, start_row + nrows + 1):
        for j in range(start_col + 1, start_col + ncols + 1):
            cell = sheet.cell(row=i, column=j)
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal='center')
    # 设置索引列的格式
    for cell in sheet["B"]:
        cell.number_format = "mmm yy"
    # 设置列名格式及宽度
    for c in range(start_col, start_col + ncols + 1):
        cell = sheet.cell(row=start_row, column=c)
        sheet.column_dimensions[cell.column_letter].width = 14

    # 画图步骤
    # 设置图属性，x轴，y轴标题，高度，宽度
    chart = BarChart()
    chart.type = 'col'
    chart.title = 'Sales Per Month and Store'
    chart.x_axis.title = summary.index.name
    chart.y_axis.title = 'Sales'
    chart.height = 22
    chart.width = 40
    # 图形引用的数据区域 data
    data = Reference(worksheet=sheet,
                     min_col=start_col + 1,
                     min_row=start_row,
                     max_col=start_col + ncols - 1,
                     max_row=start_row + nrows - 1)

    # 图形的分类 categories
    categories = Reference(worksheet=sheet,
                           min_col=start_col,
                           min_row=start_row + 1,
                           max_row=start_row + nrows - 1)
    # 作图
    chart.add_data(data=data, from_rows=False, titles_from_data=True)
    chart.set_categories(categories)
    # 设置图形的位置锚点
    cell = sheet.cell(row=start_row + nrows + 2, column=start_col)
    sheet.add_chart(chart=chart, anchor=cell.coordinate)
    chart.y_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))
